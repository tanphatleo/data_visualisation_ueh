"""
Lọc thông báo chào giá liên quan đến máy đo điện tim / Holter / ECG
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import re
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

SOURCE_FILE = "thong_bao_chao_gia_20260529_134424.xlsx"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}
DELAY = 0.8

# Từ khoá cần tìm trong nội dung bài
ECG_KEYWORDS = [
    r"holter",
    r"ecg",
    r"ekg",
    r"điện tim",
    r"dien tim",
    r"máy đo tim",
    r"may do tim",
    r"monitor tim",
    r"điện tâm đồ",
    r"dien tam do",
]
ECG_PATTERN = re.compile("|".join(ECG_KEYWORDS), re.IGNORECASE)


def extract_page_text(url: str) -> tuple[str, str, str]:
    """Fetch page, return (date_str, full_text, matched_keywords)."""
    resp = requests.get(url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "html.parser")

    # --- Date ---
    date_str = ""
    for sel in ["time", ".entry-date", ".post-date", ".published", "[class*='date']"]:
        tag = soup.select_one(sel)
        if tag:
            date_str = tag.get("datetime") or tag.get_text(strip=True)
            break

    # --- Body text (remove scripts/styles) ---
    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    text = soup.get_text(" ", strip=True)

    # --- Find matched keywords ---
    matches = set(m.group(0).lower() for m in ECG_PATTERN.finditer(text))

    return date_str, text, ", ".join(sorted(matches))


def load_thiet_bi(path: str) -> pd.DataFrame:
    df = pd.read_excel(path, index_col=0)
    mask = df["Tiêu đề"].str.contains(r"thiết bị|thiet bi", case=False, na=False)
    return df[mask].copy()


def main():
    print("Đọc file Excel và lọc 'thiết bị'...")
    df = load_thiet_bi(SOURCE_FILE)
    print(f"  → {len(df)} bài có 'thiết bị'")

    results = []
    for idx, (stt, row) in enumerate(df.iterrows(), 1):
        title = row["Tiêu đề"]
        url   = row["Đường dẫn"]
        print(f"[{idx}/{len(df)}] {title[:60]}...")

        try:
            time.sleep(DELAY)
            date_str, text, matched = extract_page_text(url)
        except Exception as e:
            print(f"  [LỖI] {e}")
            date_str, matched = "", ""

        # Check title + body for ECG keywords
        combined = title + " " + text if 'text' in dir() else title
        all_matches = set(m.group(0).lower() for m in ECG_PATTERN.finditer(combined))
        matched = ", ".join(sorted(all_matches))

        is_ecg = bool(all_matches)

        results.append({
            "STT gốc": stt,
            "Tiêu đề": title,
            "Ngày đăng": date_str,
            "Từ khóa khớp": matched,
            "Liên quan ECG/Holter": "CÓ" if is_ecg else "KHÔNG",
            "Đường dẫn": url,
        })

        if is_ecg:
            print(f"  ✓ KHỚP: {matched}")

    df_result = pd.DataFrame(results)
    df_ecg    = df_result[df_result["Liên quan ECG/Holter"] == "CÓ"]
    df_no     = df_result[df_result["Liên quan ECG/Holter"] == "KHÔNG"]

    print(f"\n=== KẾT QUẢ ===")
    print(f"Tổng bài 'thiết bị' đã quét: {len(df_result)}")
    print(f"Có liên quan ECG/Holter/Điện tim: {len(df_ecg)}")
    print(f"Không liên quan: {len(df_no)}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = f"ecg_holter_results_{timestamp}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Sheet 1: ECG matches only
        df_ecg.reset_index(drop=True).assign(STT=range(1, len(df_ecg)+1)) \
              .set_index("STT") \
              .to_excel(writer, sheet_name="Có ECG-Holter")

        # Sheet 2: All thiết bị results
        df_result.reset_index(drop=True).assign(STT=range(1, len(df_result)+1)) \
                 .set_index("STT") \
                 .to_excel(writer, sheet_name="Tất cả Thiết bị")

        # Sheet 3: Summary stats
        keyword_counts = {}
        for row in results:
            for kw in row["Từ khóa khớp"].split(", "):
                kw = kw.strip()
                if kw:
                    keyword_counts[kw] = keyword_counts.get(kw, 0) + 1

        df_stats = pd.DataFrame([
            {"Thống kê": "Tổng bài 'thiết bị'", "Số lượng": len(df_result)},
            {"Thống kê": "Có liên quan ECG/Holter/Điện tim", "Số lượng": len(df_ecg)},
            {"Thống kê": "Không liên quan", "Số lượng": len(df_no)},
        ])
        df_stats.to_excel(writer, sheet_name="Thống kê", index=False)

        if keyword_counts:
            df_kw = pd.DataFrame(
                sorted(keyword_counts.items(), key=lambda x: -x[1]),
                columns=["Từ khóa", "Số lần xuất hiện"]
            )
            df_kw.to_excel(writer, sheet_name="Thống kê", index=False, startrow=6)

    # Format Excel
    _format_excel(out_path, len(df_ecg))
    print(f"\nĐã lưu: {out_path}")
    return out_path


def _format_excel(path: str, ecg_count: int):
    wb = load_workbook(path)

    BLUE_FILL   = PatternFill("solid", fgColor="1F4E79")
    GREEN_FILL  = PatternFill("solid", fgColor="E2EFDA")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Header row formatting
        for cell in ws[1]:
            if cell.value:
                cell.fill = BLUE_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Highlight ECG rows in green (sheet 2: col F = "Liên quan ECG/Holter")
        if sheet_name == "Tất cả Thiết bị":
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 5 and cell.value == "CÓ":
                        for c in row:
                            c.fill = GREEN_FILL

        # Auto column width
        for col in ws.columns:
            max_w = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w + 2, 60)

    wb.save(path)


if __name__ == "__main__":
    main()
